import openpyxl as opxl
from openpyxl.utils.cell import get_column_letter
import os
import numpy as np
from DSF_fitting import process_single_protein

NO_OF_SAMPLES=3

# returns the separated (protein,conc) tuple
def separateProteinConc(input):
    if input[-2:] == 'uM':
        inputSplitList = input.split('_')
        protein = ('_').join(inputSplitList[0:-1])
        conc = np.float(inputSplitList[-1][:-2])
    else:
        protein=None
        conc=None
    return (protein, conc)

# returns the dictionary of temperature and rowNumber
def getTemperatureDict(workSheet):
    tempDict=dict([])
    rowIterator = 2
    currValue = workSheet['A'+str(rowIterator)].value
    while currValue is not None:
        if str(currValue)[-1] != '*':
            tempDict[np.float(currValue)]=str(rowIterator)
        rowIterator=rowIterator+1
        currValue = workSheet['A' + str(rowIterator)].value
    return tempDict


# returns a dictionary containing column number corresponding to different concs for different proteins
def getProteinDict(workSheet):
    proteinDict=dict([])
    columnIterator = 2
    columnChar=get_column_letter(columnIterator)
    currValue = workSheet[columnChar+'1'].value
    while currValue is not None:
        (protein, conc) = separateProteinConc(currValue)
        if protein is not None:
            if protein not in proteinDict.keys():
                proteinDict[protein] = dict([])
            if conc not in proteinDict[protein].keys():
                proteinDict[protein][conc]=[]
            proteinDict[protein][conc].append(columnChar)
            #proteinDict[protein] = {conc: columnChar}
        columnIterator = columnIterator + 1
        columnChar = get_column_letter(columnIterator)
        currValue = workSheet[columnChar + '1'].value
    return proteinDict

# builds the dictionary of required data
# nested dictionary key structure
# data[protein->conc->sample->temperature] = value
def buildDict(workSheet):
    data = dict([])
    tempDict = getTemperatureDict(workSheet)
    proteinDict = getProteinDict(workSheet)
    for protein in proteinDict.keys():
        for conc in proteinDict[protein].keys():
            for temperature in tempDict.keys():
                if protein not in data.keys():
                    data[protein] = dict([])
                if conc not in data[protein].keys():
                    data[protein][conc] = dict([])
                #if temperature not in data[protein][conc].keys():
                #    data[protein][conc][temperature]=dict([])
                for i in range(NO_OF_SAMPLES):
                    cellValue = workSheet[proteinDict[protein][conc][i]+tempDict[temperature]].value
                    if i not in data[protein][conc].keys():
                        data[protein][conc][i] = dict([])
                    if str(cellValue)[-1] == '*':
                        data[protein][conc][i][temperature] = None
                    else:
                        data[protein][conc][i][temperature] = np.float(cellValue)
    augmentDict(data)
    return data

# augments the data in the build Dictionary
def augmentDict(data):
    for protein in data.keys():
        for conc in data[protein].keys():
            for sample in range(NO_OF_SAMPLES):
                # determine the min and max from the list
                min_val = min(x for x in data[protein][conc][sample].values() if x is not None)
                max_val = max(x for x in data[protein][conc][sample].values() if x is not None)
                before_min = True
                after_max =  False
                for temp in data[protein][conc][sample].keys():
                    if data[protein][conc][sample][temp] == min_val:
                        before_min = False
                    if data[protein][conc][sample][temp] == max_val:
                        after_max = True
                    if before_min:
                        data[protein][conc][sample][temp] = 0.0
                    elif after_max:
                        data[protein][conc][sample][temp] = 1.0
                    else:
                        data[protein][conc][sample][temp] = (data[protein][conc][sample][temp] - min_val)/(max_val-min_val)

def generateConcFiles(protein, data, destFile):
    file=open(destFile,'w')
    concList = list(data[protein].keys())
    concList.sort()
    for conc in concList:
        for i in range(NO_OF_SAMPLES):
            file.write(str(conc)+'\n')
    file.close()

def generateFluoFiles(protein, data, destFile):
    file=open(destFile,'w')
    # write the header
    file.write("Temperature\t")
    concList = list(data[protein].keys())
    concList.sort()
    for conc in concList:
        for i in range(NO_OF_SAMPLES):
            file.write(str(conc)+'uM_'+str(i+1)+'\t')
    file.write('\n')
    for temp in data[protein][list(data[protein].keys())[0]][0].keys():
        file.write(str(temp))
        for conc in concList:
            for sample in data[protein][conc].keys():
                file.write('\t'+str(data[protein][conc][sample][temp]))
        file.write('\n')
    file.close()

def generateFiles(data, srcFile):
    # create the directory setup
    for protein in data.keys():
        if not os.path.exists('../Output/'+protein):
            os.makedirs('../Output/'+protein)
        destConcFile = '../Output/'+protein+'/'+protein+'_conc.txt'
        destFluoFile = '../Output/'+protein+'/'+protein+'_fluo.txt'
        generateConcFiles(protein, data, destConcFile)
        generateFluoFiles(protein, data, destFluoFile)

if __name__ == '__main__':
    fileList = os.listdir('../Input')
    for file in fileList:
        print('\n' + str(fileList.index(file) + 1) + ') ' + file)
    idx = int(input('Select source file : '))
    idx = idx - 1
    srcFile = fileList[idx]
    wb = opxl.load_workbook(filename='../Input/'+srcFile)
    ws = wb.active
    data = buildDict(ws)
    generateFiles(data, srcFile)
    for protein in data.keys():
        process_single_protein(protein)
