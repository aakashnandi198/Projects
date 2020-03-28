import openpyxl as opxl
import openpyxl.utils as oput
import openpyxl.styles as opst
import sys
import os

def detectColumn(ws,attribute):
    columnHeader = [x.value.lower() for x in ws[1]]
    columnNumber = columnHeader.index(attribute.lower())
    columnChar = chr(ord('A')+columnNumber)
    return columnChar

def readSrcFile(srcFile):
    # Create a dictionary to store info related to each well
    readDict={}
    # Open the excel file
    wb = opxl.load_workbook(filename=srcFile)
    ws = wb.active
    rowIterator=2
    #Automate detection
    wellPos = detectColumn(ws, 'well position')
    tempPos = detectColumn(ws, 'temperature')
    fluorePos = detectColumn(ws, 'fluorescence')
    while ws[wellPos+str(rowIterator)].value is not None:
        wellName=ws[wellPos+str(rowIterator)].value
        if not(wellName in readDict.keys()):
            readDict[wellName] = {}
        readDict[wellName][ws[tempPos+str(rowIterator)].value] = ws[fluorePos+str(rowIterator)].value
        rowIterator = rowIterator+1
    return readDict

def reformatExcel(readDict, destFile):
    wb = opxl.Workbook()
    ws = wb.active
    HeadFt = opst.Font(bold=True)
    HeadAlign = opst.Alignment(horizontal='center', vertical='center')
    flatFt = opst.Font(color='FF00FF00')
    slope_epsilon= 0.0005
    #Prepare header rows
    ws['A1'].value = 'Temperature'
    ws['A1'].font = HeadFt
    ws['A1'].alignment = HeadAlign
    ws.column_dimensions['A'].width = 20
    columnIterator = 2
    for wellpos in readDict.keys():
        # fix this
        ws[oput.get_column_letter(columnIterator)+'1'].value = wellpos
        ws[oput.get_column_letter(columnIterator) + '1'].font = HeadFt
        ws[oput.get_column_letter(columnIterator) + '1'].alignment = HeadAlign
        columnIterator = columnIterator+1

    #Preparing the temperature column
    rowIterator = 2
    firstWell = list(readDict.keys())[0]
    for temp in readDict[firstWell].keys():
        ws['A'+str(rowIterator)].value = temp
        ws['A' + str(rowIterator)].alignment = HeadAlign
        rowIterator = rowIterator+1

    columnIterator=2
    for well in readDict.keys():
        rowIterator=2
        tempList= list(readDict[well].keys())
        for temp in tempList:
            ws[oput.get_column_letter(columnIterator)+str(rowIterator)].value = readDict[well][temp]
            rowIterator = rowIterator+1
        columnIterator = columnIterator+1

    wb.save(filename=destFile)


if __name__ == '__main__':
    fileList =  os.listdir('../Input')
    for file in fileList:
        print(str(fileList.index(file)+1)+') '+file)
    idx = int(input('Select source file : '))
    idx = idx-1
    srcFile = fileList[idx]
    destFile = '../Output/Formatted_'+srcFile
    readDict = readSrcFile('../Input/'+srcFile)
    reformatExcel(readDict, destFile)
    print('\nDone')
